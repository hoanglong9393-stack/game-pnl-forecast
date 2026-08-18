import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Game P&L Forecast Pro - Hoàng Thành Long", layout="wide", page_icon="🎮")

st.title("🎮 Hệ Thống Dự Phóng P&L by Hoàng Thành Long (VplayHN)")

# ==========================================
# CUSTOM CSS FOR EXCEL-LIKE TABLE
# ==========================================
st.markdown("""
<style>
    .section-title { font-size: 18px; font-weight: 600; color: #1E293B; margin-top: 10px; margin-bottom: 8px; }
    .dataframe-container { overflow-x: auto; margin-bottom: 20px; }
    table.custom-pnl { width: 100%; border-collapse: collapse; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; font-size: 13px; color: #E2E8F0; background-color: #0F172A; }
    table.custom-pnl th { background-color: #0B3E45; color: white; font-weight: bold; text-align: center; padding: 8px 5px; border: 1px solid #1E293B; min-width: 100px; }
    table.custom-pnl th:first-child { background-color: #002B36; text-align: left; min-width: 200px; position: sticky; left: 0; z-index: 10; }
    table.custom-pnl td { padding: 6px 8px; text-align: right; border: 1px solid #334155; }
    table.custom-pnl td:first-child { text-align: left; font-weight: 500; position: sticky; left: 0; background-color: #0F172A; z-index: 9; border-right: 2px solid #475569; }
    
    table.custom-pnl tr.row-nru td { background-color: #F59E0B; color: black; font-weight: bold; }
    table.custom-pnl tr.row-dau td { background-color: #D97706; color: white; font-weight: bold; }
    table.custom-pnl tr.row-mau td { background-color: #EA580C; color: white; font-weight: bold; }
    table.custom-pnl tr.row-cost td { background-color: #FBBF24; color: black; }
    table.custom-pnl tr.row-rev-total td { background-color: #DC2626; color: white; font-weight: bold; }
    table.custom-pnl tr.row-spent-header td { background-color: #94A3B8; color: black; font-weight: bold; text-align: left;}
    table.custom-pnl tr.row-opex td { background-color: #FCD34D; color: black; }
    table.custom-pnl tr.row-total-cost td { background-color: #DC2626; color: white; font-weight: bold; }
    
    table.custom-pnl tr.row-profit-month td { background-color: white; color: black; }
    table.custom-pnl tr.row-profit-month td.positive { background-color: #22C55E; color: white; }
    table.custom-pnl tr.row-profit-cum td { background-color: white; color: black; font-weight: bold; }
    table.custom-pnl tr.row-profit-cum td.positive { background-color: #22C55E; color: white; }
    table.custom-pnl tr.row-roi td { background-color: white; color: black; }
</style>
""", unsafe_allow_html=True)

# ==========================================
# KHỞI TẠO DỮ LIỆU
# ==========================================
def get_default_fixed_costs(total_months=25):
    months_label = ["Pre-launch", "🔒 Month OB (Auto)"] + [f"Month OB+{i}" for i in range(1, total_months - 1)]
    return pd.DataFrame({
        "Tháng": months_label,
        "Nhân sự (VNĐ)": [400000000, 200000000] + [200000000] * (total_months - 2),
        "Server (VNĐ)": [0, 200000000] + [200000000] * (total_months - 2),
        "LF + Branding (VNĐ)": [675000000, 1800000000] + [50000000] * (total_months - 2)
    })

def get_default_traffic(total_months=25, is_android=True):
    months_label = ["Pre-launch", "🔒 Month OB (Auto)"] + [f"Month OB+{i}" for i in range(1, total_months - 1)]
    return pd.DataFrame({
        "Tháng": months_label,
        "NRU": [10000 if is_android else 5000, 100000 if is_android else 50000] + [70000 if is_android else 30000] * (total_months - 2),
        "CPN (VNĐ)": [15000 if is_android else 20000, 25000 if is_android else 32000] + [25000 if is_android else 32000] * (total_months - 2)
    })

def get_default_ob_daily(total_ob_nru, default_cpn):
    weights = np.zeros(30)
    weights[0:3] = [0.25, 0.15, 0.10]
    weights[3:7] = 0.05
    weights[7:30] = 0.30 / 23
    nru_days = np.round(weights * total_ob_nru).astype(int)
    nru_days[0] += total_ob_nru - np.sum(nru_days)
    return pd.DataFrame({
        "Ngày": [f"Day {i+1}" for i in range(30)],
        "NRU (Users)": nru_days,
        "CPN (VNĐ)": [default_cpn] * 30
    })

ALL_D_COLS = ["D1", "D3", "D7", "D14", "D30", "D60", "D90", "D180", "D210", "D240", "D270", "D300", "D330", "D360"]
ALL_RR_COLS = ["D1", "D3", "D7", "D14", "D30", "D60", "D90", "D180", "D360"]
ALL_D_TARGETS = [3, 7, 14, 30, 60, 90, 180, 210, 240, 270, 300, 330, 360]

def get_default_ltv(is_android=True):
    if is_android:
        return pd.DataFrame({
            "Phase Name": ["Phase 1 (Tháng OB)", "Phase 2 (Tháng 2&3)", "Phase 3 (Tháng 4+)"],
            "Áp dụng từ Tháng": ["🔒 Month OB (Auto)", "Month OB+1", "Month OB+3"],
            "D1": [9000, 7000, 4500], "D3": [13500, 10500, 7000], "D7": [27000, 22000, 13000],
            "D14": [40000, 33000, 22000], "D30": [54000, 44000, 30000], "D60": [72000, 57000, 35000],
            "D90": [85000, 66000, 40000], "D180": [108000, 80000, 45000], "D210": [112000, 83000, 46000],
            "D240": [116000, 86000, 47000], "D270": [120000, 89000, 48000], "D300": [124000, 92000, 49000],
            "D330": [128000, 95000, 50000], "D360": [132000, 98000, 51000]
        })
    else:
        return pd.DataFrame({
            "Phase Name": ["Phase 1 (Tháng OB)", "Phase 2 (Tháng 2&3)", "Phase 3 (Tháng 4+)"],
            "Áp dụng từ Tháng": ["🔒 Month OB (Auto)", "Month OB+1", "Month OB+3"],
            "D1": [12000, 10000, 6000], "D3": [18000, 15000, 10000], "D7": [36000, 31000, 19000],
            "D14": [55000, 48000, 31000], "D30": [72000, 62000, 45000], "D60": [96000, 81000, 50000],
            "D90": [115000, 93000, 55000], "D180": [144000, 110000, 60000], "D210": [150000, 113000, 61000],
            "D240": [156000, 116000, 62000], "D270": [162000, 119000, 63000], "D300": [168000, 122000, 64000],
            "D330": [174000, 125000, 65000], "D360": [180000, 128000, 66000]
        })

def get_default_rr(is_android=True):
    if is_android:
        return pd.DataFrame({
            "Phase Name": ["Phase 1 (Tháng OB)", "Phase 2 (Tháng 2&3)", "Phase 3 (Tháng 4+)"],
            "Áp dụng từ Tháng": ["🔒 Month OB (Auto)", "Month OB+1", "Month OB+3"],
            "D1": [40.0, 38.0, 35.0], "D3": [20.0, 18.0, 15.0], "D7": [10.0, 9.0, 8.0],
            "D14": [7.0, 6.0, 5.0], "D30": [4.0, 3.5, 3.0], "D60": [2.0, 1.8, 1.5],
            "D90": [1.0, 0.9, 0.8], "D180": [0.5, 0.4, 0.3], "D360": [0.1, 0.1, 0.1]
        })
    else:
        return pd.DataFrame({
            "Phase Name": ["Phase 1 (Tháng OB)", "Phase 2 (Tháng 2&3)", "Phase 3 (Tháng 4+)"],
            "Áp dụng từ Tháng": ["🔒 Month OB (Auto)", "Month OB+1", "Month OB+3"],
            "D1": [45.0, 43.0, 40.0], "D3": [25.0, 23.0, 20.0], "D7": [12.0, 11.0, 10.0],
            "D14": [8.0, 7.5, 7.0], "D30": [5.0, 4.5, 4.0], "D60": [3.0, 2.8, 2.5],
            "D90": [1.5, 1.3, 1.2], "D180": [0.8, 0.7, 0.6], "D360": [0.2, 0.2, 0.2]
        })

def highlight_ob_row(row):
    color = 'background-color: #FFCDD2; color: #B71C1C; font-weight: bold;' if row['Tháng'] == '🔒 Month OB (Auto)' else ''
    return [color] * len(row)

def migrate_month_ob(df):
    if "Tháng" in df.columns:
        df["Tháng"] = df["Tháng"].replace("Month OB", "🔒 Month OB (Auto)")
    if "Áp dụng từ Tháng" in df.columns:
        df["Áp dụng từ Tháng"] = df["Áp dụng từ Tháng"].replace("Month OB", "🔒 Month OB (Auto)")
    return df

if "project_names" not in st.session_state:
    st.session_state.project_names = ["Dự án 1 (T029)", "T037"]
    st.session_state.current_project = "T037"

# ==========================================
# SIDEBAR QUẢN LÝ DỰ ÁN & NỀN TẢNG
# ==========================================
with st.sidebar:
    st.header("📁 Quản Lý Dự Án")
    selected_proj = st.selectbox("Chọn dự án:", st.session_state.project_names, index=st.session_state.project_names.index(st.session_state.current_project))
    st.session_state.current_project = selected_proj
    cur_proj = st.session_state.current_project
    
    if f"platforms_{cur_proj}" not in st.session_state:
        st.session_state[f"platforms_{cur_proj}"] = ["Android", "iOS"]
        
    current_platforms = st.session_state[f"platforms_{cur_proj}"]
    
    with st.expander("➕ Tạo Dự Án Mới"):
        new_proj_name = st.text_input("Tên dự án mới:")
        new_proj_months = st.number_input("Số tháng dự phóng", min_value=3, max_value=60, value=25)
        if st.button("Tạo & Lưu"):
            if new_proj_name and new_proj_name not in st.session_state.project_names:
                st.session_state.project_names.append(new_proj_name)
                st.session_state[f"platforms_{new_proj_name}"] = ["Android", "iOS"]
                st.session_state[f"fixed_costs_{new_proj_name}"] = get_default_fixed_costs(new_proj_months)
                st.session_state[f"traffic_Android_{new_proj_name}"] = get_default_traffic(new_proj_months, True)
                st.session_state[f"traffic_iOS_{new_proj_name}"] = get_default_traffic(new_proj_months, False)
                st.session_state[f"ob_daily_Android_{new_proj_name}"] = get_default_ob_daily(100000, 25000)
                st.session_state[f"ob_daily_iOS_{new_proj_name}"] = get_default_ob_daily(50000, 32000)
                st.session_state[f"ltv_Android_{new_proj_name}"] = get_default_ltv(True)
                st.session_state[f"ltv_iOS_{new_proj_name}"] = get_default_ltv(False)
                st.session_state[f"rr_Android_{new_proj_name}"] = get_default_rr(True)
                st.session_state[f"rr_iOS_{new_proj_name}"] = get_default_rr(False)
                st.session_state[f"params_{new_proj_name}"] = {"rev_share": 20.2, "vat": 10.0, "payment_fee": 5.0, "prelaunch_comeback_pct": 60.0, "usd_rate": 25400.0}
                st.session_state.current_project = new_proj_name
                st.rerun()

    with st.expander("🗑️ Xóa Dự Án Hiện Tại"):
        if len(st.session_state.project_names) > 1:
            st.warning(f"Chắc chắn xóa dự án **{cur_proj}**?")
            if st.button("⚠️ Xác nhận Xóa", type="primary", use_container_width=True):
                st.session_state.project_names.remove(cur_proj)
                keys_to_delete = [k for k in st.session_state.keys() if k.endswith(f"_{cur_proj}")]
                for k in keys_to_delete: del st.session_state[k]
                st.session_state.current_project = st.session_state.project_names[0]
                st.success("Đã xóa dự án thành công!")
                st.rerun()
        else:
            st.info("Không thể xóa dự án duy nhất.")

    st.markdown("---")
    st.header("📱 Quản Lý Nền Tảng (Sources)")
    with st.expander("➕ Thêm Nền Tảng Mới"):
        new_plat = st.text_input("Tên nền tảng (VD: Web, PC, Huawei):")
        if st.button("Thêm Nền Tảng"):
            if new_plat and new_plat not in current_platforms:
                st.session_state[f"platforms_{cur_proj}"].append(new_plat)
                months_len = len(st.session_state.get(f"fixed_costs_{cur_proj}", get_default_fixed_costs(25)))
                st.session_state[f"traffic_{new_plat}_{cur_proj}"] = get_default_traffic(months_len, False)
                st.session_state[f"ob_daily_{new_plat}_{cur_proj}"] = get_default_ob_daily(20000, 15000)
                st.session_state[f"ltv_{new_plat}_{cur_proj}"] = get_default_ltv(False)
                st.session_state[f"rr_{new_plat}_{cur_proj}"] = get_default_rr(False)
                st.success(f"Đã thêm nền tảng {new_plat}!")
                st.rerun()

    st.markdown("---")
    st.header("📤 Tải Lên Dữ Liệu (Excel)")
    uploaded_file = st.file_uploader("Upload file PNL_*.xlsx", type=["xlsx"])
    
    if uploaded_file is not None:
        if st.session_state.get("last_uploaded_file_id") != uploaded_file.file_id:
            with st.spinner("Đang đọc file Excel & Tương thích ngược..."):
                try:
                    excel_data = pd.ExcelFile(uploaded_file)
                    imported_proj_name = uploaded_file.name.replace("PNL_", "").replace("_Input", "").replace("_Report", "").replace(".xlsx", "")
                    if imported_proj_name not in st.session_state.project_names:
                        st.session_state.project_names.append(imported_proj_name)
                    
                    found_plats = []
                    for sheet in excel_data.sheet_names:
                        if sheet.startswith("Traffic "): found_plats.append(sheet.replace("Traffic ", ""))
                    if not found_plats: found_plats = ["Android", "iOS"]
                    st.session_state[f"platforms_{imported_proj_name}"] = found_plats
                    
                    if 'Fixed Costs' in excel_data.sheet_names:
                        st.session_state[f"fixed_costs_{imported_proj_name}"] = migrate_month_ob(pd.read_excel(excel_data, sheet_name='Fixed Costs').fillna(0))
                    elif 'Traffic Android' in excel_data.sheet_names:
                        old_tr = pd.read_excel(excel_data, sheet_name='Traffic Android').fillna(0)
                        if "Nhân sự (VNĐ)" in old_tr.columns:
                            fc_cols = ["Tháng"] + [c for c in ["Nhân sự (VNĐ)", "Server (VNĐ)", "LF + Branding (VNĐ)"] if c in old_tr.columns]
                            st.session_state[f"fixed_costs_{imported_proj_name}"] = migrate_month_ob(old_tr[fc_cols])
                        else:
                            st.session_state[f"fixed_costs_{imported_proj_name}"] = get_default_fixed_costs(len(old_tr))
                            
                    for p in found_plats:
                        if f'Traffic {p}' in excel_data.sheet_names:
                            tr_df = pd.read_excel(excel_data, sheet_name=f'Traffic {p}').fillna(0)
                            clean_cols = [c for c in tr_df.columns if c not in ["Nhân sự (VNĐ)", "Server (VNĐ)", "LF + Branding (VNĐ)"]]
                            st.session_state[f"traffic_{p}_{imported_proj_name}"] = migrate_month_ob(tr_df[clean_cols])
                        if f'OB Daily {p}' in excel_data.sheet_names:
                            st.session_state[f"ob_daily_{p}_{imported_proj_name}"] = pd.read_excel(excel_data, sheet_name=f'OB Daily {p}').fillna(0)
                        if f'LTV {p}' in excel_data.sheet_names:
                            st.session_state[f"ltv_{p}_{imported_proj_name}"] = migrate_month_ob(pd.read_excel(excel_data, sheet_name=f'LTV {p}').fillna(0))
                        
                        if f'RR {p}' in excel_data.sheet_names:
                            st.session_state[f"rr_{p}_{imported_proj_name}"] = migrate_month_ob(pd.read_excel(excel_data, sheet_name=f'RR {p}').fillna(0))
                        else:
                            st.session_state[f"rr_{p}_{imported_proj_name}"] = get_default_rr(p=="Android")
                            
                    if 'Params' in excel_data.sheet_names:
                        st.session_state[f"params_{imported_proj_name}"] = pd.read_excel(excel_data, sheet_name='Params').to_dict('records')[0]
                        
                    st.session_state.current_project = imported_proj_name
                    st.session_state["last_uploaded_file_id"] = uploaded_file.file_id
                    st.success(f"Đã nạp thành công dự án '{imported_proj_name}'!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Lỗi đọc file Excel: {e}")

    st.markdown("---")
    st.header("⚙️ Cấu Hình Tham Số Chung")
    if f"params_{cur_proj}" not in st.session_state:
        st.session_state[f"params_{cur_proj}"] = {"rev_share": 20.2, "vat": 10.0, "payment_fee": 5.0, "prelaunch_comeback_pct": 60.0, "usd_rate": 25400.0}
    p_params = st.session_state[f"params_{cur_proj}"]
    
    usd_rate = st.number_input("Tỷ giá USD/VNĐ", value=float(p_params.get("usd_rate", 25400.0)), step=100.0)
    rev_share_pct = st.number_input("Revenue Share Dev (%)", value=float(p_params.get("rev_share", 20.2)), step=0.1)
    vat_pct = st.number_input("VAT (%)", value=float(p_params.get("vat", 10.0)), step=0.5)
    payment_fee_pct = st.number_input("Payment Fee (%)", value=float(p_params.get("payment_fee", 5.0)), step=0.5)
    st.markdown("👉 **Chỉ số User chuyển đổi**")
    prelaunch_comeback_pct = st.number_input("Tỷ lệ Pre-launch quay lại ngày OB (%)", value=float(p_params.get("prelaunch_comeback_pct", 60.0)), step=1.0)
    
    st.session_state[f"params_{cur_proj}"] = {
        "usd_rate": usd_rate,
        "rev_share": rev_share_pct, "vat": vat_pct, "payment_fee": payment_fee_pct,
        "prelaunch_comeback_pct": prelaunch_comeback_pct
    }

# ==========================================
# KHỞI TẠO STATE CHO DỰ ÁN HIỆN TẠI (NẾU CHƯA CÓ)
# ==========================================
if f"fixed_costs_{cur_proj}" not in st.session_state: st.session_state[f"fixed_costs_{cur_proj}"] = get_default_fixed_costs(25)
for p in current_platforms:
    if f"traffic_{p}_{cur_proj}" not in st.session_state: st.session_state[f"traffic_{p}_{cur_proj}"] = get_default_traffic(25, p=="Android")
    if f"ob_daily_{p}_{cur_proj}" not in st.session_state: st.session_state[f"ob_daily_{p}_{cur_proj}"] = get_default_ob_daily(100000 if p=="Android" else 50000, 25000 if p=="Android" else 32000)
    if f"ltv_{p}_{cur_proj}" not in st.session_state: st.session_state[f"ltv_{p}_{cur_proj}"] = get_default_ltv(p=="Android")
    if f"rr_{p}_{cur_proj}" not in st.session_state: st.session_state[f"rr_{p}_{cur_proj}"] = get_default_rr(p=="Android")

# ==========================================
# CƯỠNG CHẾ ĐỒNG BỘ DỮ LIỆU MONTH OB TỪ BẢNG DAILY
# ==========================================
for p in current_platforms:
    current_tr = st.session_state[f"traffic_{p}_{cur_proj}"]
    ob_daily = st.session_state[f"ob_daily_{p}_{cur_proj}"]
    params = st.session_state[f"params_{cur_proj}"]

    ob_daily_nru_sum = float(ob_daily["NRU (Users)"].sum())
    ob_daily_budget = float((ob_daily["NRU (Users)"] * ob_daily["CPN (VNĐ)"]).sum())

    # TÍNH NGẦM TRÊN FINAL P&L, HIỂN THỊ CHỈ HIỂN THỊ NRU MỚI KHÔNG CỘNG DỒN COMEBACK
    total_ob_nru = int(np.round(ob_daily_nru_sum))
    calc_ob_cpn = int(np.round(ob_daily_budget / total_ob_nru)) if total_ob_nru > 0 else 0

    idx_ob = current_tr[current_tr["Tháng"] == "🔒 Month OB (Auto)"].index
    if len(idx_ob) > 0:
        current_tr.loc[idx_ob[0], "NRU"] = total_ob_nru
        current_tr.loc[idx_ob[0], "CPN (VNĐ)"] = calc_ob_cpn
    st.session_state[f"traffic_{p}_{cur_proj}"] = current_tr

# ==========================================
# TABS HIỂN THỊ CHÍNH
# ==========================================
tabs_names = ["💸 Chi Phí Cố Định"] + [f"📱 {p}" for p in current_platforms] + ["📊 Báo Cáo P&L Tổng Hợp"]
rendered_tabs = st.tabs(tabs_names)

# TAB 0: CHI PHÍ CỐ ĐỊNH
with rendered_tabs[0]:
    st.markdown(f'<div class="section-title">💸 Kế Hoạch Chi Phí Cố Định Khác (Fixed Costs) - {cur_proj}</div>', unsafe_allow_html=True)
    st.info("Bảng này chứa các chi phí không phụ thuộc trực tiếp vào số lượng User (Server, Nhân sự, LF, Branding).")
    
    fc_df = st.session_state[f"fixed_costs_{cur_proj}"]
    styled_fc = fc_df.style.apply(highlight_ob_row, axis=1)
    
    edited_fc = st.data_editor(
        styled_fc, num_rows="dynamic", use_container_width=True, hide_index=True, key=f"ed_fc_{cur_proj}",
        column_config={
            "Nhân sự (VNĐ)": st.column_config.NumberColumn("Nhân sự (VNĐ)", format="%d", min_value=0),
            "Server (VNĐ)": st.column_config.NumberColumn("Server (VNĐ)", format="%d", min_value=0),
            "LF + Branding (VNĐ)": st.column_config.NumberColumn("LF + Branding (VNĐ)", format="%d", min_value=0)
        }
    )
    if not edited_fc.astype(str).equals(fc_df.astype(str)):
        st.session_state[f"fixed_costs_{cur_proj}"] = edited_fc
        st.rerun()

# CÁC TAB NỀN TẢNG (TRAFFIC, RR & LTV)
for idx, p in enumerate(current_platforms):
    with rendered_tabs[idx + 1]:
        col_title, col_btn = st.columns([4, 1])
        col_title.markdown(f'<div class="section-title">Nền Tảng: {p} ({cur_proj})</div>', unsafe_allow_html=True)
        if len(current_platforms) > 1:
            if col_btn.button(f"🗑️ Xóa {p}", key=f"del_{p}_{cur_proj}"):
                st.session_state[f"platforms_{cur_proj}"].remove(p)
                st.rerun()

        st.markdown(f"**1. Kế Hoạch Traffic Tháng - {p}**")
        st.info("🔒 Dòng bôi đỏ `Month OB` được tự động tính toán từ bảng phân bổ 30 ngày. Vui lòng không ghi đè.")
        
        df_p = st.session_state[f"traffic_{p}_{cur_proj}"]
        styled_p = df_p.style.apply(highlight_ob_row, axis=1)
        
        edited_tr = st.data_editor(
            styled_p, num_rows="dynamic", use_container_width=True, hide_index=True, key=f"ed_tr_{p}_{cur_proj}",
            column_config={
                "NRU": st.column_config.NumberColumn(f"NRU {p}", format="%d", min_value=0),
                "CPN (VNĐ)": st.column_config.NumberColumn(f"CPN {p} (VNĐ)", format="%d", min_value=0)
            }
        )
        if not edited_tr.astype(str).equals(df_p.astype(str)):
            st.session_state[f"traffic_{p}_{cur_proj}"] = edited_tr
            st.rerun()
            
        with st.expander(f"📅 Chi Tiết Phân Bổ 30 Ngày Tháng OPEN BETA ({p}) - [Sửa số tại đây]", expanded=False):
            ob_df = st.session_state[f"ob_daily_{p}_{cur_proj}"]
            cur_sum_nru = float(ob_df["NRU (Users)"].sum())
            cur_sum_budget = float((ob_df["NRU (Users)"] * ob_df["CPN (VNĐ)"]).sum())
            
            c1, c2, c3, c4 = st.columns([1.5, 1.5, 2, 2])
            target_nru = c1.number_input(f"Mục tiêu NRU mua ({p}):", value=int(cur_sum_nru) if cur_sum_nru > 0 else 50000, step=1000, key=f"tgt_nru_{p}_{cur_proj}")
            target_cpn = c2.number_input(f"Mục tiêu CPN mua TB ({p}):", value=int(cur_sum_budget/cur_sum_nru) if cur_sum_nru > 0 else 25000, step=1000, key=f"tgt_cpn_{p}_{cur_proj}")
            with c3:
                st.write(""); st.write("")
                if st.button("⚡ Chia mốc dồn đầu (50-20-30)", key=f"btn_dist_{p}_{cur_proj}"):
                    st.session_state[f"ob_daily_{p}_{cur_proj}"] = get_default_ob_daily(target_nru, target_cpn)
                    st.rerun()
                    
            edited_ob = st.data_editor(
                st.session_state[f"ob_daily_{p}_{cur_proj}"], num_rows="fixed", use_container_width=True, hide_index=True, key=f"ed_ob_{p}_{cur_proj}",
                column_config={"NRU (Users)": st.column_config.NumberColumn("NRU", format="%d", min_value=0), "CPN (VNĐ)": st.column_config.NumberColumn("CPN", format="%d", min_value=0)}
            )
            if not edited_ob.astype(str).equals(st.session_state[f"ob_daily_{p}_{cur_proj}"].astype(str)):
                st.session_state[f"ob_daily_{p}_{cur_proj}"] = edited_ob
                st.rerun()

        st.markdown("---")
        st.markdown(f"**2. Cấu Hình Retention Rate (RR %) - {p}**")
        month_options = df_p["Tháng"].tolist()
        
        col_cfg_rr = {"Áp dụng từ Tháng": st.column_config.SelectboxColumn("Áp dụng từ Tháng", options=month_options)}
        for c in ALL_RR_COLS: col_cfg_rr[c] = st.column_config.NumberColumn(f"{c} (%)", format="%.2f", min_value=0.0, max_value=100.0)
        
        edited_rr = st.data_editor(
            st.session_state[f"rr_{p}_{cur_proj}"], num_rows="dynamic", use_container_width=True, hide_index=True, column_config=col_cfg_rr, key=f"ed_rr_{p}_{cur_proj}"
        )
        if not edited_rr.astype(str).equals(st.session_state[f"rr_{p}_{cur_proj}"].astype(str)):
            st.session_state[f"rr_{p}_{cur_proj}"] = edited_rr

        st.markdown("---")
        st.markdown(f"**3. Cấu Hình LTV Curve & Hệ Số K - {p}**")
        col_cfg = {"Áp dụng từ Tháng": st.column_config.SelectboxColumn("Áp dụng từ Tháng", options=month_options)}
        for c in ALL_D_COLS: col_cfg[c] = st.column_config.NumberColumn(f"{c} (VNĐ)", format="%d", min_value=0)
        
        edited_ltv = st.data_editor(
            st.session_state[f"ltv_{p}_{cur_proj}"], num_rows="dynamic", use_container_width=True, hide_index=True, column_config=col_cfg, key=f"ed_ltv_{p}_{cur_proj}"
        )
        if not edited_ltv.astype(str).equals(st.session_state[f"ltv_{p}_{cur_proj}"].astype(str)):
            st.session_state[f"ltv_{p}_{cur_proj}"] = edited_ltv
            
        k_df = edited_ltv.copy()
        for c in ALL_D_COLS: k_df[c] = pd.to_numeric(k_df[c], errors="coerce").fillna(0.0)
        k_df['K1'] = np.where(k_df['D1'] > 0, 1.0, 0.0)
        for d in ALL_D_TARGETS: k_df[f'K{d}'] = np.where(k_df['D1'] > 0, k_df[f'D{d}'] / k_df['D1'], 0.0)
        st.dataframe(k_df[["Phase Name", "Áp dụng từ Tháng", "K1"] + [f'K{d}' for d in ALL_D_TARGETS]].style.format({f'K{d}': "{:.2f}x" for d in [1] + ALL_D_TARGETS}), use_container_width=True, hide_index=True)

# ==========================================
# ENGINE CALCULATION & EXPORT HELPERS
# ==========================================
def create_daily_ltv_curve(anchor_points):
    days = sorted(anchor_points.keys())
    max_day = 720
    full_curve = np.zeros(max_day + 1)
    for i in range(len(days) - 1):
        d_start, d_end = days[i], days[i+1]
        v_start, v_end = anchor_points[d_start], anchor_points[d_end]
        full_curve[d_start:d_end+1] = np.linspace(v_start, v_end, d_end - d_start + 1)
    last_day = days[-1]
    last_val = anchor_points[last_day]
    full_curve[last_day:max_day+1] = last_val
    return full_curve

def create_daily_rr_curve(anchor_points):
    days = sorted(anchor_points.keys())
    max_day = 720
    full_curve = np.zeros(max_day + 1)
    for i in range(len(days) - 1):
        d_start, d_end = days[i], days[i+1]
        v_start, v_end = anchor_points[d_start] / 100.0, anchor_points[d_end] / 100.0
        full_curve[d_start:d_end+1] = np.linspace(v_start, v_end, d_end - d_start + 1)
    last_day = days[-1]
    last_val = anchor_points[last_day] / 100.0
    full_curve[last_day:max_day+1] = last_val
    return full_curve

def calculate_platform_dau_phase_mapping(df_traffic, df_ob_daily, df_rr):
    num_months = len(df_traffic)
    rr_mapping = {}
    for _, row in df_rr.iterrows():
        try:
            anchors = {0: 100.0} # D0 RR is 100%
            for c in ALL_RR_COLS:
                if c in row and not pd.isna(row[c]):
                    anchors[int(c[1:])] = float(row[c])
            curve = create_daily_rr_curve(anchors)
            if row['Áp dụng từ Tháng'] in df_traffic['Tháng'].values:
                rr_mapping[row['Áp dụng từ Tháng']] = curve
        except: pass
        
    latest_curve = np.zeros(721)
    active_curve = latest_curve
    month_curves = []
    for m in df_traffic['Tháng']:
        if m in rr_mapping: active_curve = rr_mapping[m]
        month_curves.append(active_curve)
        
    daily_nru_list = []
    daily_curve_list = []
    
    for m_idx, row in df_traffic.iterrows():
        if row['Tháng'] == "🔒 Month OB (Auto)":
            ob_nrus = df_ob_daily["NRU (Users)"].astype(float).values
            for d in range(30):
                daily_nru_list.append(ob_nrus[d] if d < len(ob_nrus) else 0.0)
                daily_curve_list.append(month_curves[m_idx])
        else:
            nru_daily = float(row['NRU']) / 30.0
            for d in range(30):
                daily_nru_list.append(nru_daily)
                daily_curve_list.append(month_curves[m_idx])
                
    total_days = len(daily_nru_list)
    daily_dau = np.zeros(total_days)
    
    for c_day in range(total_days):
        c_nru = daily_nru_list[c_day]
        if c_nru <= 0: continue
        c_curve = daily_curve_list[c_day]
        for age in range(min(len(c_curve), total_days - c_day)):
            daily_dau[c_day + age] += c_nru * c_curve[age]
            
    peak_dau_arr = np.array([np.max(daily_dau[i*30:(i+1)*30]) for i in range(num_months)])
    
    mau_arr = np.zeros(num_months)
    for m in range(num_months):
        nru_m = np.sum(daily_nru_list[m*30:(m+1)*30])
        mau_m = nru_m
        mid_month_day = m*30 + 15
        for past_d in range(m*30):
            c_nru = daily_nru_list[past_d]
            if c_nru <= 0: continue
            c_curve = daily_curve_list[past_d]
            age = mid_month_day - past_d
            if age < len(c_curve):
                mr = min(c_curve[age] * 2.0, 1.0)
                mau_m += c_nru * mr
        mau_arr[m] = mau_m
        
    return peak_dau_arr, mau_arr

def calculate_platform_rev_phase_mapping(df_traffic, df_ob_daily, df_ltv):
    num_months = len(df_traffic)
    ltv_mapping = {}
    for _, row in df_ltv.iterrows():
        try:
            anchors = {int(c[1:]): float(row[c]) for c in ALL_D_COLS if c in row}
            curve = create_daily_ltv_curve(anchors)
            if row['Áp dụng từ Tháng'] in df_traffic['Tháng'].values:
                ltv_mapping[row['Áp dụng từ Tháng']] = curve
        except: pass
        
    latest_curve = np.zeros(721)
    active_curve = latest_curve
    month_curves = []
    for m in df_traffic['Tháng']:
        if m in ltv_mapping: active_curve = ltv_mapping[m]
        month_curves.append(active_curve)
        
    daily_nru_list = []
    daily_curve_list = []
    
    for m_idx, row in df_traffic.iterrows():
        if row['Tháng'] == "🔒 Month OB (Auto)":
            ob_nrus = df_ob_daily["NRU (Users)"].astype(float).values
            for d in range(30):
                daily_nru_list.append(ob_nrus[d] if d < len(ob_nrus) else 0.0)
                daily_curve_list.append(month_curves[m_idx])
        else:
            nru_daily = float(row['NRU']) / 30.0
            for d in range(30):
                daily_nru_list.append(nru_daily)
                daily_curve_list.append(month_curves[m_idx])
                
    total_days = len(daily_nru_list)
    daily_rev = np.zeros(total_days)
    for c_day in range(total_days):
        c_nru = daily_nru_list[c_day]
        if c_nru <= 0: continue
        c_curve = daily_curve_list[c_day]
        inc_ltv = np.diff(np.insert(c_curve, 0, 0))[1:]
        for age in range(min(len(inc_ltv), total_days - c_day)):
            daily_rev[c_day + age] += c_nru * inc_ltv[age]
            
    return np.array([np.sum(daily_rev[i*30:(i+1)*30]) for i in range(num_months)])

def format_cell_value(val, is_pct=False, is_usd=False):
    if pd.isna(val) or val == 0: return "0"
    if is_pct: return f"{val*100:.2f}%"
    if is_usd: return f"${val:,.2f}"
    return f"{val:,.0f}"

def format_pnl_for_excel(res):
    metrics = [
        'NRU', 'Peak DAU', 'MAU', 'CPN', 'Revenue', 
        'Nhân sự', 'Server', 'Marketing (UA+Tax)', 'LF + Branding',
        'Revenue share dev', 'VAT', 'Payment channel fee',
        'Tổng Chi Phí', 'Lợi nhuận tháng', 'Lợi Nhuận', 'Tỷ Trọng MKT/REV'
    ]
    df_export = pd.DataFrame({"Dashboard": metrics})
    
    totals = []
    for m in metrics:
        if m == 'NRU': totals.append(res['NRU'].sum())
        elif m == 'Peak DAU': totals.append(res['Peak DAU'].max())
        elif m == 'MAU': totals.append(res['MAU'].max())
        elif m == 'CPN': 
            t_nru = res['NRU'].sum()
            t_mkt = res['Marketing (UA+Tax)'].sum() + res['LF + Branding'].sum()
            totals.append(t_mkt / t_nru if t_nru > 0 else 0)
        elif m == 'Revenue': totals.append(res['Revenue'].sum())
        elif m in ['Nhân sự', 'Server', 'Marketing (UA+Tax)', 'LF + Branding', 'Revenue share dev', 'VAT', 'Payment channel fee', 'Tổng Chi Phí']:
            totals.append(res[m].sum())
        elif m == 'Lợi Nhuận': totals.append(res['Lợi Nhuận'].iloc[-1])
        elif m == 'Tỷ Trọng MKT/REV': 
            t_rev = res['Revenue'].sum()
            t_mkt = res['Marketing (UA+Tax)'].sum()
            totals.append(t_mkt / t_rev if t_rev > 0 else 0)
        else: totals.append(None)
        
    df_export["Total"] = totals
    for _, row in res.iterrows():
        df_export[row['Tháng']] = [row.get(m, None) for m in metrics]
        
    df_with_kpi = pd.DataFrame([[""] + ["KPI"] * (len(df_export.columns) - 1)], columns=df_export.columns)
    df_export = pd.concat([df_with_kpi, df_export], ignore_index=True)
    
    df_export['Dashboard'] = df_export['Dashboard'].replace('NRU', 'New Registed User')
    return df_export

def generate_report_excel(res_vnd, res_usd, current_platforms, cur_proj):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_vnd = format_pnl_for_excel(res_vnd)
        df_usd = format_pnl_for_excel(res_usd)
        
        df_vnd.to_excel(writer, sheet_name='P&L Tổng Hợp (VNĐ)', index=False)
        df_usd.to_excel(writer, sheet_name='P&L (USD)', index=False)
        
        for p in current_platforms:
            tr = st.session_state[f"traffic_{p}_{cur_proj}"].copy()
            tr["Tháng"] = tr["Tháng"].replace("🔒 Month OB (Auto)", "Month OB")
            ltv = st.session_state[f"ltv_{p}_{cur_proj}"].copy()
            rr_exp = st.session_state[f"rr_{p}_{cur_proj}"].copy()
            
            if "Áp dụng từ Tháng" in ltv.columns:
                ltv["Áp dụng từ Tháng"] = ltv["Áp dụng từ Tháng"].replace("🔒 Month OB (Auto)", "Month OB")
            if "Áp dụng từ Tháng" in rr_exp.columns:
                rr_exp["Áp dụng từ Tháng"] = rr_exp["Áp dụng từ Tháng"].replace("🔒 Month OB (Auto)", "Month OB")
            
            tr.to_excel(writer, sheet_name=p, startrow=1, index=False)
            ws = writer.sheets[p]
            ws.cell(row=1, column=1, value=f"KẾ HOẠCH TRAFFIC - {p}").font = Font(bold=True)
            
            r_idx_rr = len(tr) + 4
            ws.cell(row=r_idx_rr, column=1, value=f"CẤU HÌNH RETENTION RATE (RR %) - {p}").font = Font(bold=True)
            rr_exp.to_excel(writer, sheet_name=p, startrow=r_idx_rr, index=False)
            
            r_idx_ltv = r_idx_rr + len(rr_exp) + 3
            ws.cell(row=r_idx_ltv, column=1, value=f"CẤU HÌNH LTV - {p}").font = Font(bold=True)
            ltv.to_excel(writer, sheet_name=p, startrow=r_idx_ltv, index=False)

        fill_header = PatternFill(start_color="0B3E45", end_color="0B3E45", fill_type="solid")
        fill_nru = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        fill_dau = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        fill_mau = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
        fill_cpn = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
        fill_rev_cost = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        fill_spent = PatternFill(start_color="94A3B8", end_color="94A3B8", fill_type="solid")
        fill_opex = PatternFill(start_color="FCD34D", end_color="FCD34D", fill_type="solid")
        fill_profit_pos = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        font_white_bold = Font(color="FFFFFF", bold=True)
        font_black_bold = Font(color="000000", bold=True)
        font_black = Font(color="000000")
        font_white = Font(color="FFFFFF")
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for sheet_name in ['P&L Tổng Hợp (VNĐ)', 'P&L (USD)']:
            ws = writer.book[sheet_name]
            
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_white_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                row_label = str(row[0].value)
                current_fill = fill_white
                current_font = font_black
                
                if row_label == 'New Registed User':
                    current_fill = fill_nru
                    current_font = font_black_bold
                elif row_label == 'Peak DAU':
                    current_fill = fill_dau
                    current_font = font_white_bold
                elif row_label == 'MAU':
                    current_fill = fill_mau
                    current_font = font_white_bold
                elif row_label == 'CPN':
                    current_fill = fill_cpn
                elif row_label in ['Revenue', 'Tổng Chi Phí']:
                    current_fill = fill_rev_cost
                    current_font = font_white_bold
                elif row_label == 'Spent':
                    current_fill = fill_spent
                    current_font = font_black_bold
                elif row_label in ['Nhân sự', 'Server', 'Marketing (UA+Tax)', 'LF + Branding', 'Revenue share dev', 'VAT', 'Payment channel fee']:
                    current_fill = fill_opex
                elif row_label == 'Lợi Nhuận':
                    current_font = font_black_bold
                    
                for idx, cell in enumerate(row):
                    cell_font = current_font
                    cell_fill = current_fill
                    
                    if row_label in ['Lợi nhuận tháng', 'Lợi Nhuận'] and idx > 0:
                        if isinstance(cell.value, (int, float)) and cell.value > 0:
                            cell_fill = fill_profit_pos
                            cell_font = font_white if row_label == 'Lợi nhuận tháng' else font_white_bold
                            
                    cell.fill = cell_fill
                    cell.font = cell_font
                    cell.border = thin_border
                    
                    ws.freeze_panes = 'C3'
                    
                    if idx > 0 and isinstance(cell.value, (int, float)):
                        if "Tỷ Trọng" in row_label:
                            cell.number_format = '0.00%'
                        elif sheet_name == 'P&L (USD)':
                            cell.number_format = '"$"#,##0.00'
                        else:
                            cell.number_format = '#,##0'

        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[col_letter].width = min(max_length + 3, 25)
                
    buffer.seek(0)
    return buffer

def generate_pnl_html(res, is_usd=False):
    total_nru = res['NRU'].sum()
    total_mkt = res['Marketing (UA+Tax)'].sum()
    total_lf = res['LF + Branding'].sum()
    cpn_total = (total_mkt + total_lf) / total_nru if total_nru > 0 else 0

    html = '<div class="dataframe-container"><table class="custom-pnl">'
    html += '<tr><th>Dashboard</th><th>Total</th>'
    for m in res['Tháng']: html += f'<th>{m}</th>'
    html += '</tr>'
    
    html += '<tr><td></td><td>KPI</td>'
    for _ in res['Tháng']: html += '<td>KPI</td>'
    html += '</tr>'
    
    html += f'<tr class="row-nru"><td>New Registed User</td><td>{format_cell_value(total_nru)}</td>'
    for v in res['NRU']: html += f'<td>{format_cell_value(v)}</td>'
    html += '</tr>'

    html += f'<tr class="row-dau"><td>Peak DAU</td><td>{format_cell_value(res["Peak DAU"].max())}</td>'
    for v in res['Peak DAU']: html += f'<td>{format_cell_value(v)}</td>'
    html += '</tr>'
    
    html += f'<tr class="row-mau"><td>MAU</td><td>{format_cell_value(res["MAU"].max())}</td>'
    for v in res['MAU']: html += f'<td>{format_cell_value(v)}</td>'
    html += '</tr>'
    
    html += f'<tr class="row-cost"><td>CPN</td><td>{format_cell_value(cpn_total, is_usd=is_usd)}</td>'
    for v in res['CPN']: html += f'<td>{format_cell_value(v, is_usd=is_usd)}</td>'
    html += '</tr>'
    
    html += f'<tr class="row-rev-total"><td>Revenue</td><td>{format_cell_value(res["Revenue"].sum(), is_usd=is_usd)}</td>'
    for v in res['Revenue']: html += f'<td>{format_cell_value(v, is_usd=is_usd)}</td>'
    html += '</tr>'
    
    html += '<tr class="row-spent-header"><td>Spent</td><td></td>' + '<td></td>'*len(res) + '</tr>'
    
    opex_rows = [
        ('Personel', 'Nhân sự'), ('Server', 'Server'),
        ('Marketing (UA+Tax)', 'Marketing (UA+Tax)'), ('LF + Branding', 'LF + Branding'),
        ('Revenue share dev', 'Revenue share dev'), ('VAT', 'VAT'),
        ('Payment channel fee', 'Payment channel fee')
    ]
    for label, col in opex_rows:
        html += f'<tr class="row-opex"><td>{label}</td><td>{format_cell_value(res[col].sum(), is_usd=is_usd)}</td>'
        for v in res[col]: html += f'<td>{format_cell_value(v, is_usd=is_usd)}</td>'
        html += '</tr>'
        
    html += f'<tr class="row-total-cost"><td>Tổng Chi Phí</td><td>{format_cell_value(res["Tổng Chi Phí"].sum(), is_usd=is_usd)}</td>'
    for v in res['Tổng Chi Phí']: html += f'<td>{format_cell_value(v, is_usd=is_usd)}</td>'
    html += '</tr>'
    
    html += '<tr class="row-profit-month"><td>Lợi nhuận tháng</td><td></td>'
    for v in res['Lợi nhuận tháng']: 
        cls = "positive" if v > 0 else ""
        html += f'<td class="{cls}">{format_cell_value(v, is_usd=is_usd)}</td>'
    html += '</tr>'
    
    total_profit = res['Lợi Nhuận'].iloc[-1]
    html += f'<tr class="row-profit-cum"><td>Lợi Nhuận</td><td>{format_cell_value(total_profit, is_usd=is_usd)}</td>'
    for v in res['Lợi Nhuận']: 
        cls = "positive" if v > 0 else ""
        html += f'<td class="{cls}">{format_cell_value(v, is_usd=is_usd)}</td>'
    html += '</tr>'
    
    avg_mkt_rev = total_mkt / res["Revenue"].sum() if res["Revenue"].sum() > 0 else 0
    html += f'<tr class="row-roi"><td>Tỷ Trọng MKT/REV</td><td>{format_cell_value(avg_mkt_rev, is_pct=True)}</td>'
    for v in res['Tỷ Trọng MKT/REV']: html += f'<td>{format_cell_value(v, is_pct=True)}</td>'
    html += '</tr>'
    
    html += '</table></div>'
    return html

# TAB CUỐI CÙNG: BÁO CÁO P&L TỔNG HỢP
with rendered_tabs[-1]:
    st.markdown(f'<div class="section-title">📊 Báo Cáo P&L Tổng Hợp (Consolidated) - {cur_proj}</div>', unsafe_allow_html=True)
    
    params = st.session_state[f"params_{cur_proj}"]
    fixed_costs = st.session_state[f"fixed_costs_{cur_proj}"]
    
    run_sim = st.button(f"🚀 Chạy Mô Phỏng Tổng Đa Nền Tảng ({cur_proj})", type="primary")
    
    if run_sim or f"pnl_res_{cur_proj}" in st.session_state:
        with st.spinner("Đang tính ma trận Cohort hợp nhất đa nền tảng..."):
            res = pd.DataFrame()
            display_months = fixed_costs['Tháng'].replace("🔒 Month OB (Auto)", "Month OB").tolist()
            res['Tháng'] = display_months
            
            total_nru_arr = np.zeros(len(res))
            total_peak_dau_arr = np.zeros(len(res))
            total_mau_arr = np.zeros(len(res))
            total_mkt_arr = np.zeros(len(res))
            total_rev_arr = np.zeros(len(res))
            comeback_rate = params.get('prelaunch_comeback_pct', 60.0) / 100.0
            
            for p in current_platforms:
                tr_df = st.session_state[f"traffic_{p}_{cur_proj}"]
                ob_df = st.session_state[f"ob_daily_{p}_{cur_proj}"]
                ltv_df = st.session_state[f"ltv_{p}_{cur_proj}"]
                rr_df = st.session_state[f"rr_{p}_{cur_proj}"]
                
                pre_nru = float(tr_df.loc[tr_df['Tháng'] == 'Pre-launch', 'NRU'].sum())
                ob_calc = ob_df.copy()
                if len(ob_calc) > 0:
                    ob_calc.loc[0, "NRU (Users)"] += pre_nru * comeback_rate
                
                p_nru, p_mkt = [], []
                for _, r in tr_df.iterrows():
                    if r['Tháng'] == "🔒 Month OB (Auto)":
                        p_nru.append(ob_df["NRU (Users)"].sum())  # TRÁNH DOUBLE COUNT TRÊN BẢNG P&L
                        p_mkt.append((ob_df["NRU (Users)"] * ob_df["CPN (VNĐ)"]).sum())
                    else:
                        u = float(r['NRU'])
                        p_nru.append(u)
                        p_mkt.append(u * float(r['CPN (VNĐ)']))
                        
                total_nru_arr += np.array(p_nru)
                total_mkt_arr += np.array(p_mkt)
                total_rev_arr += calculate_platform_rev_phase_mapping(tr_df, ob_calc, ltv_df)
                
                p_peak_dau, p_mau = calculate_platform_dau_phase_mapping(tr_df, ob_calc, rr_df)
                total_peak_dau_arr += p_peak_dau
                total_mau_arr += p_mau
                
            res['NRU'] = total_nru_arr
            res['Peak DAU'] = total_peak_dau_arr
            res['MAU'] = total_mau_arr
            res['Marketing (UA+Tax)'] = total_mkt_arr
            res['Revenue'] = total_rev_arr
            
            res['Nhân sự'] = pd.to_numeric(fixed_costs['Nhân sự (VNĐ)'], errors='coerce').fillna(0.0).values
            res['Server'] = pd.to_numeric(fixed_costs['Server (VNĐ)'], errors='coerce').fillna(0.0).values
            res['LF + Branding'] = pd.to_numeric(fixed_costs['LF + Branding (VNĐ)'], errors='coerce').fillna(0.0).values
            
            res['CPN'] = np.where(res['NRU'] > 0, (res['Marketing (UA+Tax)'] + res['LF + Branding']) / res['NRU'], 0.0)
            res['Revenue share dev'] = res['Revenue'] * (params.get('rev_share', 20.2) / 100.0)
            res['VAT'] = res['Revenue'] * (params.get('vat', 10.0) / 100.0)
            res['Payment channel fee'] = res['Revenue'] * (params.get('payment_fee', 5.0) / 100.0)
            
            res['Tổng Chi Phí'] = (
                res['Marketing (UA+Tax)'] + res['Nhân sự'] + res['Server'] +
                res['LF + Branding'] + res['Revenue share dev'] + res['VAT'] + res['Payment channel fee']
            )
            res['Lợi nhuận tháng'] = res['Revenue'] - res['Tổng Chi Phí']
            res['Lợi Nhuận'] = res['Lợi nhuận tháng'].cumsum()
            res['Tỷ Trọng MKT/REV'] = np.where(res['Revenue'] > 0, res['Marketing (UA+Tax)'] / res['Revenue'], 0.0)
            st.session_state[f"pnl_res_{cur_proj}"] = res
            
            res_usd = res.copy()
            usd_rate = float(params.get('usd_rate', 25400.0))
            monetary_cols = ['Marketing (UA+Tax)', 'Revenue', 'Nhân sự', 'Server', 'LF + Branding', 'CPN', 'Revenue share dev', 'VAT', 'Payment channel fee', 'Tổng Chi Phí', 'Lợi nhuận tháng', 'Lợi Nhuận']
            for col in monetary_cols: res_usd[col] = res_usd[col] / usd_rate
            
            col_down1, col_down2 = st.columns(2)
            
            buffer_input = io.BytesIO()
            with pd.ExcelWriter(buffer_input, engine='openpyxl') as writer:
                fc_export = fixed_costs.copy()
                fc_export["Tháng"] = fc_export["Tháng"].replace("🔒 Month OB (Auto)", "Month OB")
                fc_export.to_excel(writer, sheet_name='Fixed Costs', index=False)
                
                for p in current_platforms:
                    tr_exp = st.session_state[f"traffic_{p}_{cur_proj}"].copy()
                    tr_exp["Tháng"] = tr_exp["Tháng"].replace("🔒 Month OB (Auto)", "Month OB")
                    ltv_exp = st.session_state[f"ltv_{p}_{cur_proj}"].copy()
                    rr_exp = st.session_state[f"rr_{p}_{cur_proj}"].copy()
                    
                    if "Áp dụng từ Tháng" in ltv_exp.columns:
                        ltv_exp["Áp dụng từ Tháng"] = ltv_exp["Áp dụng từ Tháng"].replace("🔒 Month OB (Auto)", "Month OB")
                    if "Áp dụng từ Tháng" in rr_exp.columns:
                        rr_exp["Áp dụng từ Tháng"] = rr_exp["Áp dụng từ Tháng"].replace("🔒 Month OB (Auto)", "Month OB")
                        
                    tr_exp.to_excel(writer, sheet_name=f'Traffic {p}', index=False)
                    st.session_state[f"ob_daily_{p}_{cur_proj}"].to_excel(writer, sheet_name=f'OB Daily {p}', index=False)
                    rr_exp.to_excel(writer, sheet_name=f'RR {p}', index=False)
                    ltv_exp.to_excel(writer, sheet_name=f'LTV {p}', index=False)
                pd.DataFrame([params]).to_excel(writer, sheet_name='Params', index=False)
            buffer_input.seek(0)
            
            col_down1.download_button(
                label=f"📥 Tải File Cấu Hình Input (Dùng để upload lại)", data=buffer_input,
                file_name=f"PNL_{cur_proj}_Input.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            buffer_report = generate_report_excel(res, res_usd, current_platforms, cur_proj)
            col_down2.download_button(
                label=f"📊 Tải Báo Cáo P&L & Cấu Hình (Excel)", data=buffer_report,
                file_name=f"PNL_{cur_proj}_Report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.markdown("### 🇻🇳 Báo Cáo P&L (VNĐ)")
            st.markdown(generate_pnl_html(res, is_usd=False), unsafe_allow_html=True)
            
            st.markdown(f"### 🇺🇸 Báo Cáo P&L (USD) - *Tỷ giá: {usd_rate:,.0f} đ*")
            st.markdown(generate_pnl_html(res_usd, is_usd=True), unsafe_allow_html=True)
